import pandas as pd
import pytest

from rag_lib.loaders.csv_excel import ExcelLoader


def _require_openpyxl() -> None:
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        pytest.skip("openpyxl not installed")


def test_excel_loader_csv_mode_returns_one_doc_per_sheet(tmp_path):
    _require_openpyxl()

    xlsx_path = tmp_path / "multisheet.xlsx"
    with pd.ExcelWriter(xlsx_path) as writer:
        pd.DataFrame({"col1": [1, 2], "col2": ["a", "b"]}).to_excel(
            writer, sheet_name="Sheet1", index=False
        )
        pd.DataFrame({"name": ["foo"], "value": [42]}).to_excel(
            writer, sheet_name="Sheet2", index=False
        )

    loader = ExcelLoader(str(xlsx_path), output_format="csv")
    docs = loader.load()

    assert len(docs) == 2
    assert docs[0].metadata["sheet_name"] == "Sheet1"
    assert docs[1].metadata["sheet_name"] == "Sheet2"
    assert docs[0].metadata["output_format"] == "csv"
    assert docs[1].metadata["output_format"] == "csv"
    assert "col1,col2" in docs[0].page_content
    assert "name,value" in docs[1].page_content


def test_excel_loader_markdown_mode_keeps_text_and_multiple_tables(tmp_path):
    _require_openpyxl()
    from openpyxl import Workbook

    xlsx_path = tmp_path / "mixed.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"] = "Quarterly report notes"
    ws["A3"] = "Product"
    ws["B3"] = "Revenue"
    ws["A4"] = "A"
    ws["B4"] = 10

    ws["A6"] = "Comment: seasonal uplift"
    ws["A8"] = "Region"
    ws["B8"] = "Share"
    ws["A9"] = "EU"
    ws["B9"] = "40%"

    wb.save(xlsx_path)

    loader = ExcelLoader(str(xlsx_path), output_format="markdown")
    docs = loader.load()

    assert len(docs) == 1
    doc = docs[0]
    assert doc.metadata["sheet_name"] == "Sheet1"
    assert doc.metadata["output_format"] == "markdown"

    # Text blocks are preserved.
    assert "Quarterly report notes" in doc.page_content
    assert "Comment: seasonal uplift" in doc.page_content

    # Multiple table blocks are preserved in one sheet document.
    assert "| Product" in doc.page_content
    assert "Revenue" in doc.page_content
    assert "| Region" in doc.page_content
    assert "Share" in doc.page_content
